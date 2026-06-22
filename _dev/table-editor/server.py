import os
import csv
import glob
from flask import Flask, jsonify, request, send_from_directory
import openpyxl

app = Flask(__name__)

EXCELS_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../Assets/Resources/Excels'))


def get_all_sheets():
    result = []
    files = sorted(glob.glob(os.path.join(EXCELS_DIR, '*.xlsx')))
    for path in files:
        fname = os.path.splitext(os.path.basename(path))[0]
        if fname.startswith('~$'):
            continue
        wb = openpyxl.load_workbook(path, read_only=True)
        for sheet_name in wb.sheetnames:
            result.append({'file': fname, 'sheet': sheet_name})
        wb.close()

    # CSV 파일도 포함
    for path in sorted(glob.glob(os.path.join(EXCELS_DIR, '*.csv'))):
        fname = os.path.splitext(os.path.basename(path))[0]
        result.append({'file': fname, 'sheet': fname, 'csv': True})

    return result


def load_xlsx_sheet(file_name, sheet_name):
    path = os.path.join(EXCELS_DIR, file_name + '.xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    headers = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val is None:
            break
        headers.append(str(val))

    rows = []
    for r in range(2, ws.max_row + 1):
        first = ws.cell(r, 1).value
        if first is None:
            break
        if str(first).startswith('#'):
            continue
        row = {}
        for i, h in enumerate(headers):
            val = ws.cell(r, i + 1).value
            row[h] = val if val is not None else ''
        rows.append(row)

    wb.close()
    return headers, rows


def save_xlsx_sheet(file_name, sheet_name, headers, rows):
    path = os.path.join(EXCELS_DIR, file_name + '.xlsx')
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]

    # 헤더 쓰기
    for i, h in enumerate(headers):
        ws.cell(1, i + 1, h)

    # 기존 데이터 행 지우기
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(r, c, None)

    # 새 데이터 쓰기
    for ri, row in enumerate(rows):
        for ci, h in enumerate(headers):
            val = row.get(h, '')
            if isinstance(val, str) and val != '':
                try:
                    val = int(val)
                except ValueError:
                    try:
                        val = float(val)
                    except ValueError:
                        pass
            ws.cell(ri + 2, ci + 1, val if val != '' else None)

    wb.save(path)


def load_csv_sheet(file_name):
    path = os.path.join(EXCELS_DIR, file_name + '.csv')
    with open(path, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        all_rows = list(reader)

    if not all_rows:
        return [], []

    headers = [h for h in all_rows[0] if h]
    rows = []
    for r in all_rows[1:]:
        if not r or not r[0] or r[0].startswith('#'):
            continue
        row = {}
        for i, h in enumerate(headers):
            row[h] = r[i] if i < len(r) else ''
        rows.append(row)

    return headers, rows


def save_csv_sheet(file_name, headers, rows):
    path = os.path.join(EXCELS_DIR, file_name + '.csv')
    with open(path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([row.get(h, '') for h in headers])


@app.route('/')
def index():
    return send_from_directory(os.path.dirname(__file__), 'index.html')


@app.route('/api/tables')
def api_tables():
    return jsonify(get_all_sheets())


@app.route('/api/table/<file_name>/<sheet_name>')
def api_table(file_name, sheet_name):
    try:
        csv_path = os.path.join(EXCELS_DIR, file_name + '.csv')
        if os.path.exists(csv_path):
            headers, rows = load_csv_sheet(file_name)
        else:
            headers, rows = load_xlsx_sheet(file_name, sheet_name)
        return jsonify({'headers': headers, 'rows': rows})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/table/<file_name>/<sheet_name>', methods=['POST'])
def api_table_save(file_name, sheet_name):
    try:
        data = request.json
        csv_path = os.path.join(EXCELS_DIR, file_name + '.csv')
        if os.path.exists(csv_path):
            save_csv_sheet(file_name, data['headers'], data['rows'])
        else:
            save_xlsx_sheet(file_name, sheet_name, data['headers'], data['rows'])
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    print(f'Excels dir: {EXCELS_DIR}')
    app.run(debug=True, port=5000)
